"""
Build Cardinal Health conference cohort dashboard JSON and Clifford enrollment Excel.

Merges relabel + scored CSVs, writes outputs/cardinal_cohort_dashboard.json and
outputs/Cardinal_Health_Longitudinal_Enrollment_Lists.xlsx for the Reference Cohort UI tab.
"""

from __future__ import annotations

import json
from pathlib import Path

import numpy as np
import pandas as pd

BASE = Path(__file__).resolve().parents[2]
GOOD_ONES = BASE / "Diabetes-KiHealth" / "TL-KiHealth" / "Good-Ones-Kihealth"
OUTPUTS = BASE / "outputs"

RELABEL_PATH = GOOD_ONES / "cardinal_health_2026_relabel.csv"
SCORED_PATH = OUTPUTS / "cardinal_health_2026_scored.csv"
JSON_PATH = OUTPUTS / "cardinal_cohort_dashboard.json"
UI_JSON_PATH = BASE / "kihealth_ui" / "cardinal_cohort_dashboard.json"
EXCEL_PATH = OUTPUTS / "Cardinal_Health_Longitudinal_Enrollment_Lists.xlsx"

FOLLOW_UP_THRESHOLD = 10.0
HIGH_SIGNAL_THRESHOLD = 15.0

NOTES_AVG = {
    "182943": "INS 399 = 0, signal driven by other sites",
    "501386": "INS 399 = 2.8, signal driven by other sites",
}

EXCEL_COLUMNS = [
    "UIN",
    "Age",
    "Gender",
    "BMI",
    "A1c",
    "INS 399 %",
    "3-Site Average %",
    "Cascade Result",
    "Diabetes Self-Report",
    "T1D Self-Report",
    "T2D Self-Report",
    "Clinical Note",
]

ORANGE_TOP_N = 8
ORANGE_FILL = "FFDAB9"


def is_yes_like(val) -> bool:
    if pd.isna(val):
        return False
    text = str(val).strip().lower()
    return text not in {"", "no", "n", "nan", "none", "nat"}


def fmt_report_val(val) -> str:
    if pd.isna(val):
        return ""
    return str(val).strip()


def report_conflicts(row: pd.Series) -> list[str]:
    conflicts: list[str] = []
    main_yes = row.get("diabetes_diagnosed") == "Yes"
    main_no = row.get("diabetes_diagnosed") == "No"
    type_yes = is_yes_like(row.get("q_t1d_date")) or is_yes_like(row.get("q_t2d_date"))
    hba1c = row.get("hba1c_percent")

    if main_yes and not type_yes:
        conflicts.append("Diabetes=Yes but T1/T2 dates all No")
    if main_no and type_yes:
        parts = []
        if is_yes_like(row.get("q_t1d_date")):
            parts.append(f"T1D={fmt_report_val(row.get('q_t1d_date'))}")
        if is_yes_like(row.get("q_t2d_date")):
            parts.append(f"T2D={fmt_report_val(row.get('q_t2d_date'))}")
        conflicts.append("Diabetes=No but " + ", ".join(parts))
    if main_no and pd.notna(hba1c) and float(hba1c) >= 5.7:
        conflicts.append(f"Diabetes=No but A1c {float(hba1c):.2f}")
    if main_yes and pd.notna(hba1c) and float(hba1c) < 5.7:
        conflicts.append(f"Diabetes=Yes but A1c normal ({float(hba1c):.2f})")
    return conflicts


def is_clean_diagnosed_yes(row: pd.Series) -> bool:
    return row.get("diabetes_diagnosed") == "Yes" and not report_conflicts(row)


def follow_up_mask(df: pd.DataFrame) -> pd.Series:
    """Enrollment lists: unascertained, undiagnosed dysglycemia, or conflicting self-report."""
    has_conflict = df.apply(lambda row: bool(report_conflicts(row)), axis=1)
    base = (df["unascertained"] == 1) | (df["undiagnosed_dysglycemia"] == 1) | has_conflict
    exclude = df.apply(is_clean_diagnosed_yes, axis=1)
    return base & ~exclude


def clinical_note(row: pd.Series, *, list_kind: str = "399") -> str:
    conflicts = report_conflicts(row)
    donor_id = str(row["donor_id"])
    hba1c = row.get("hba1c_percent")
    ins_399 = row.get("ins_399_pct_unmeth")

    if conflicts:
        conflict_txt = "; ".join(conflicts)
        if row.get("undiagnosed_dysglycemia") == 1 and pd.notna(hba1c):
            a1c = float(hba1c)
            if a1c >= 6.5:
                return f"URGENT: Undiagnosed diabetes — {conflict_txt}"
            if a1c >= 5.7:
                return f"URGENT: Undiagnosed prediabetes — {conflict_txt}"
        return f"Self-report conflict — {conflict_txt}"

    if donor_id == "109384":
        return "High INS 399 with normal A1c — clinical follow-up recommended"
    if donor_id in NOTES_AVG and list_kind == "avg":
        return NOTES_AVG[donor_id]
    if pd.notna(ins_399) and float(ins_399) >= FOLLOW_UP_THRESHOLD:
        return "Early signal: A1c normal, high INS 399"
    return "Early signal: A1c normal, high 3-site average"


def note_399(row: pd.Series) -> str:
    return clinical_note(row, list_kind="399")


def note_avg(row: pd.Series) -> str:
    return clinical_note(row, list_kind="avg")


def build_follow_up_row(row: pd.Series, note: str) -> dict:
    conflicts = report_conflicts(row)
    return {
        "donor_id": row["donor_id"],
        "age": round(float(row["age_years"]), 1) if pd.notna(row["age_years"]) else None,
        "gender": str(row["gender"]) if pd.notna(row["gender"]) else "",
        "hba1c": round(float(row["hba1c_percent"]), 2) if pd.notna(row["hba1c_percent"]) else None,
        "ins_399": round(float(row["ins_399_pct_unmeth"]), 2)
        if pd.notna(row["ins_399_pct_unmeth"])
        else None,
        "average_3site": round(float(row["beta_score_average"]), 2)
        if pd.notna(row["beta_score_average"])
        else None,
        "cascade": str(row["cascade_result"]) if pd.notna(row["cascade_result"]) else "",
        "diabetes_self_report": fmt_report_val(row.get("diabetes_diagnosed")),
        "t1d_self_report": fmt_report_val(row.get("q_t1d_date")),
        "t2d_self_report": fmt_report_val(row.get("q_t2d_date")),
        "self_report_conflicts": conflicts,
        "note": note,
    }


def build_self_report_record(row: pd.Series, *, excluded: bool = False) -> dict:
    conflicts = report_conflicts(row)
    note_parts = []
    if excluded:
        note_parts.append("Already diagnosed (consistent Yes) — not for longitudinal enrollment")
    if conflicts:
        note_parts.append("; ".join(conflicts))
    return {
        "donor_id": row["donor_id"],
        "diabetes_self_report": fmt_report_val(row.get("diabetes_diagnosed")),
        "t1d_self_report": fmt_report_val(row.get("q_t1d_date")),
        "t2d_self_report": fmt_report_val(row.get("q_t2d_date")),
        "hba1c": round(float(row["hba1c_percent"]), 2) if pd.notna(row["hba1c_percent"]) else None,
        "ins_399": round(float(row["ins_399_pct_unmeth"]), 2)
        if pd.notna(row["ins_399_pct_unmeth"])
        else None,
        "average_3site": round(float(row["beta_score_average"]), 2)
        if pd.notna(row["beta_score_average"])
        else None,
        "conflicts": conflicts,
        "excluded_from_enrollment": excluded,
        "note": " | ".join(note_parts) if note_parts else "",
    }


def fmt_id(val) -> str:
    if pd.isna(val):
        return ""
    try:
        num = float(val)
        if num.is_integer():
            return str(int(num))
    except (TypeError, ValueError):
        pass
    return str(val).strip()


def stratum_key(hba1c: float) -> str | None:
    if pd.isna(hba1c):
        return None
    if hba1c >= 6.5:
        return "diabetic_a1c"
    if hba1c >= 5.7:
        return "prediabetes_a1c"
    return "normal_a1c"


def chart_stratum_label(hba1c: float) -> str | None:
    if pd.isna(hba1c):
        return None
    if hba1c >= 6.5:
        return "Diabetic 6.5+"
    if hba1c >= 5.7:
        return "Prediabetes 5.7-6.49"
    if hba1c >= 5.5:
        return "High-Normal 5.5-5.69"
    return "Normal <5.5"


def stats_block(series: pd.Series) -> dict:
    s = pd.to_numeric(series, errors="coerce").dropna()
    if s.empty:
        return {"mean": None, "median": None, "n": 0}
    return {"mean": round(float(s.mean()), 2), "median": round(float(s.median()), 2), "n": int(len(s))}


def to_excel_frame(df: pd.DataFrame, note_fn) -> pd.DataFrame:
    rows = []
    for _, row in df.iterrows():
        note = note_fn(row)
        rows.append(
            {
                "UIN": row["donor_id"],
                "Age": round(float(row["age_years"]), 1) if pd.notna(row["age_years"]) else None,
                "Gender": row["gender"],
                "BMI": round(float(row["bmi_kg_m2"]), 1) if pd.notna(row["bmi_kg_m2"]) else None,
                "A1c": round(float(row["hba1c_percent"]), 2) if pd.notna(row["hba1c_percent"]) else None,
                "INS 399 %": round(float(row["ins_399_pct_unmeth"]), 2)
                if pd.notna(row["ins_399_pct_unmeth"])
                else None,
                "3-Site Average %": round(float(row["beta_score_average"]), 2)
                if pd.notna(row["beta_score_average"])
                else None,
                "Cascade Result": row["cascade_result"],
                "Diabetes Self-Report": row["diabetes_diagnosed"],
                "T1D Self-Report": fmt_report_val(row.get("q_t1d_date")),
                "T2D Self-Report": fmt_report_val(row.get("q_t2d_date")),
                "Clinical Note": note,
            }
        )
    return pd.DataFrame(rows, columns=EXCEL_COLUMNS)


def confirmed_mask(df: pd.DataFrame) -> pd.Series:
    return (df["lab_dysglycemia"] == 1) | (df["at_risk_confident"] == 1)


def confirmed_note(row: pd.Series) -> str:
    if row.get("undiagnosed_dysglycemia") == 1:
        return f"Undiagnosed dysglycemia (self-report No, A1c {row['hba1c_percent']:.2f})"
    if row.get("ascertained_diabetes") == 1:
        return "Ascertained diabetes — record only (not for enrollment)"
    return "Lab dysglycemia / at-risk confirmed — record only"


def write_excel(
    ins399_df: pd.DataFrame,
    avg_df: pd.DataFrame,
    confirmed_df: pd.DataFrame,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill

    wb = Workbook()
    orange = PatternFill(start_color=ORANGE_FILL, end_color=ORANGE_FILL, fill_type="solid")

    def write_sheet(ws, frame: pd.DataFrame, highlight_top: int = 0) -> None:
        ws.append(list(frame.columns))
        for row in frame.itertuples(index=False):
            ws.append(list(row))
        if highlight_top:
            for row_idx in range(2, min(2 + highlight_top, ws.max_row + 1)):
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = orange

    wb.remove(wb.active)

    ws1 = wb.create_sheet("INS 399 Priority")
    excel_399 = to_excel_frame(ins399_df, note_399)
    write_sheet(ws1, excel_399, highlight_top=ORANGE_TOP_N)

    ws2 = wb.create_sheet("3-Site Average Priority")
    excel_avg = to_excel_frame(avg_df, note_avg)
    write_sheet(ws2, excel_avg)

    ws3 = wb.create_sheet("Confirmed Cases")
    confirmed_rows = []
    for _, row in confirmed_df.iterrows():
        confirmed_rows.append(
            {
                "UIN": row["donor_id"],
                "Age": round(float(row["age_years"]), 1) if pd.notna(row["age_years"]) else None,
                "Gender": row["gender"],
                "BMI": round(float(row["bmi_kg_m2"]), 1) if pd.notna(row["bmi_kg_m2"]) else None,
                "A1c": round(float(row["hba1c_percent"]), 2) if pd.notna(row["hba1c_percent"]) else None,
                "INS 399 %": round(float(row["ins_399_pct_unmeth"]), 2)
                if pd.notna(row["ins_399_pct_unmeth"])
                else None,
                "3-Site Average %": round(float(row["beta_score_average"]), 2)
                if pd.notna(row["beta_score_average"])
                else None,
                "Cascade Result": row["cascade_result"],
                "Diabetes Self-Report": row["diabetes_diagnosed"],
                "T1D Self-Report": fmt_report_val(row.get("q_t1d_date")),
                "T2D Self-Report": fmt_report_val(row.get("q_t2d_date")),
                "Clinical Note": confirmed_note(row),
            }
        )
    write_sheet(ws3, pd.DataFrame(confirmed_rows, columns=EXCEL_COLUMNS))

    OUTPUTS.mkdir(parents=True, exist_ok=True)
    wb.save(EXCEL_PATH)


def main() -> None:
    relabel = pd.read_csv(RELABEL_PATH)
    relabel["donor_id"] = relabel["donor_id"].map(fmt_id)

    scored = pd.read_csv(SCORED_PATH)
    scored["donor_id"] = scored["donor_id"].map(fmt_id)
    drop_cols = [c for c in scored.columns if c in relabel.columns and c != "donor_id"]
    scored_slim = scored.drop(columns=drop_cols, errors="ignore")

    df = relabel.merge(scored_slim, on="donor_id", how="left", suffixes=("", "_scored"))

    fu = follow_up_mask(df)
    ins399_list = df.loc[fu & (df["ins_399_pct_unmeth"] >= FOLLOW_UP_THRESHOLD)].sort_values(
        "ins_399_pct_unmeth", ascending=False
    )
    avg_list = df.loc[fu & (df["beta_score_average"] >= FOLLOW_UP_THRESHOLD)].sort_values(
        "beta_score_average", ascending=False
    )
    confirmed = df.loc[confirmed_mask(df)].sort_values("hba1c_percent", ascending=False)

    cascade_counts = df["cascade_result"].value_counts()
    cascade_distribution = {
        "High Confidence": int(cascade_counts.get("High Confidence", 0)),
        "Moderate": int(cascade_counts.get("Moderate", 0)),
        "Low-Moderate": int(cascade_counts.get("Low-Moderate", 0)),
        "Cleared": int(cascade_counts.get("Cleared", 0)),
    }

    ins_399_by_stratum = {
        "normal_a1c": stats_block(df.loc[df["hba1c_percent"].apply(lambda x: stratum_key(x) == "normal_a1c"), "ins_399_pct_unmeth"]),
        "prediabetes_a1c": stats_block(
            df.loc[df["hba1c_percent"].apply(lambda x: stratum_key(x) == "prediabetes_a1c"), "ins_399_pct_unmeth"]
        ),
        "diabetic_a1c": stats_block(
            df.loc[df["hba1c_percent"].apply(lambda x: stratum_key(x) == "diabetic_a1c"), "ins_399_pct_unmeth"]
        ),
    }

    chart_strata = {}
    for label in ["Normal <5.5", "High-Normal 5.5-5.69", "Prediabetes 5.7-6.49", "Diabetic 6.5+"]:
        mask = df["hba1c_percent"].apply(lambda x, lab=label: chart_stratum_label(x) == lab)
        subset = df.loc[mask]
        ins_block = stats_block(subset["ins_399_pct_unmeth"])
        avg_block = stats_block(subset["beta_score_average"])
        chart_strata[label] = {
            "n": ins_block["n"],
            "ins_399": {"mean": ins_block["mean"], "median": ins_block["median"]},
            "average_3site": {"mean": avg_block["mean"], "median": avg_block["median"]},
        }

    follow_up_list_399 = [
        build_follow_up_row(row, note_399(row)) for _, row in ins399_list.iterrows()
    ]
    follow_up_list_average = [
        build_follow_up_row(row, note_avg(row)) for _, row in avg_list.iterrows()
    ]

    has_conflict = df.apply(lambda row: bool(report_conflicts(row)), axis=1)
    clean_yes = df.apply(is_clean_diagnosed_yes, axis=1)
    high_signal = (df["ins_399_pct_unmeth"] >= FOLLOW_UP_THRESHOLD) | (
        df["beta_score_average"] >= FOLLOW_UP_THRESHOLD
    )
    self_report_conflicts = [
        build_self_report_record(row)
        for _, row in df.loc[has_conflict].iterrows()
    ]
    excluded_from_enrollment = [
        build_self_report_record(row, excluded=True)
        for _, row in df.loc[clean_yes].iterrows()
    ]
    excluded_high_signal = [
        build_self_report_record(row, excluded=True)
        for _, row in df.loc[clean_yes & high_signal].iterrows()
    ]

    pct_cf = pd.to_numeric(df["pct_cfDNA"], errors="coerce")
    payload = {
        "summary": {
            "total_samples": int(len(df)),
            "pct_cfDNA_mean": round(float(pct_cf.mean()), 2) if pct_cf.notna().any() else None,
            "collection_date": "July 2026",
            "collection_context": "Conference / walk-up screening (non-fasting)",
            "ascertained_diabetes": int((df["ascertained_diabetes"] == 1).sum()),
            "undiagnosed_dysglycemia": int((df["undiagnosed_dysglycemia"] == 1).sum()),
            "unascertained_high_signal": int(
                ((df["unascertained"] == 1) & (df["ins_399_pct_unmeth"] >= HIGH_SIGNAL_THRESHOLD)).sum()
            ),
            "cleared": int(cascade_distribution["Cleared"]),
            "diabetes_self_report_yes": int((df["diabetes_diagnosed"] == "Yes").sum()),
            "self_report_conflicts": int(has_conflict.sum()),
            "excluded_clean_yes_enrollment": int(clean_yes.sum()),
        },
        "cascade_distribution": cascade_distribution,
        "follow_up_list_399": follow_up_list_399,
        "follow_up_list_average": follow_up_list_average,
        "self_report_conflicts": self_report_conflicts,
        "excluded_from_enrollment": excluded_from_enrollment,
        "excluded_high_signal": excluded_high_signal,
        "ins_399_by_stratum": ins_399_by_stratum,
        "ins_399_chart_strata": chart_strata,
    }

    OUTPUTS.mkdir(parents=True, exist_ok=True)
    payload_text = json.dumps(payload, indent=2) + "\n"
    JSON_PATH.write_text(payload_text)
    UI_JSON_PATH.write_text(payload_text)
    write_excel(ins399_list, avg_list, confirmed)

    print(f"Saved: {JSON_PATH}")
    print(f"Saved: {UI_JSON_PATH}")
    print(f"Saved: {EXCEL_PATH}")
    print(f"INS 399 priority: {len(follow_up_list_399)} | 3-site average: {len(follow_up_list_average)} | confirmed: {len(confirmed)}")


if __name__ == "__main__":
    main()
